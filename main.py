import os
import json
import time
import platform
import subprocess
from typing import Any
from pathlib import Path
from datetime import datetime, timedelta

# requests
# https://requests.readthedocs.io/en/latest/
import requests

# openpyxl
# https://openpyxl.readthedocs.io/en/stable/
from openpyxl import Workbook
from openpyxl.styles import Border, Side, PatternFill, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


JPI_ASCII_ART = """
       _ _____ _____   ______                       _            
      | |  __ \_   _| |  ____|                     | |           
      | | |__) || |   | |__  __  ___ __   ___  _ __| |_ ___ _ __ 
  _   | |  ___/ | |   |  __| \ \/ / '_ \ / _ \| '__| __/ _ \ '__|
 | |__| | |    _| |_  | |____ >  <| |_) | (_) | |  | ||  __/ |   
  \____/|_|   |_____| |______/_/\_\ .__/ \___/|_|   \__\___|_|   
                                  | |                            
                                  |_|                            
"""

JPI_API_KEY = "688ec158-7550-42d6-abf0-3eceb2af664a"
JPI_TIME_FORMAT = "%Y-%m-%dT%H:%M:%S"
JPI_JOBS_URL = "https://api.just-plan-it.com/v1/jobs"
JPI_SETTINGS_URL = "https://api.just-plan-it.com/v1/settings"
JPI_RESOURCES_URL = "https://api.just-plan-it.com/v1/resources"
JPI_ACTIVE_JOBS = ["Planned", "Started"]
JPI_ACTIVE_TASKS = ["Planned", "Started"]
JPI_MACHINE_GROUP = "5345152d-48f0-47aa-b86a-300e0442da3a"
JPI_EXPORT_FILENAME = "JPI Machine Schedule.xlsx"


def get_jpi_settings() -> dict[str, Any]:
    # settings = json.loads(Path("settings.json").read_text())
    headers = {
        "accept": "application/json",
        "X-Api-Key": JPI_API_KEY,
    }
    response = requests.get(JPI_SETTINGS_URL, headers=headers)
    assert response.ok, "Unable to fetch settings from JPI."
    settings = response.json()
    assert isinstance(settings, dict), "Received invalid response from JPI."
    return settings


def check_job_active(job: dict[str, Any]) -> bool:
    assert isinstance(job, dict), "Invalid job data."
    return job["ExecuteStatus"] in JPI_ACTIVE_JOBS


def check_task_active(task: dict[str, Any]) -> bool:
    assert isinstance(task, dict), "Invalid task data."
    return task["TaskStatus"] in JPI_ACTIVE_TASKS


def get_active_jobs() -> list[dict[str, Any]]:
    # jobs = json.loads(Path("jobs.json").read_text())
    headers = {
        "accept": "application/json",
        "X-Api-Key": JPI_API_KEY,
    }
    response = requests.get(JPI_JOBS_URL, headers=headers)
    assert response.ok, "Unable to fetch jobs from JPI."
    jobs = response.json()
    assert isinstance(jobs, list), "Received invalid response from JPI."
    active_jobs = list(filter(check_job_active, jobs))
    for job in active_jobs:
        assert isinstance(job, dict), "Invalid job data."
        tasks = job["Tasks"]
        assert isinstance(tasks, list), "Invalid tasks data."
        active_tasks = list(filter(check_task_active, tasks))
        job["Tasks"] = active_tasks
    return active_jobs


def check_resource_group(resource: dict[str, Any]) -> bool:
    assert isinstance(resource, dict), "Invalid resource data."
    groups = resource["ResourceGroups"]
    assert isinstance(groups, list), "Invalid resources data."
    for group in groups:
        if group["Guid"] == JPI_MACHINE_GROUP:
            return True
    return False


def get_machine_resources() -> list[dict[str, Any]]:
    # resources = json.loads(Path("resources.json").read_text())
    headers = {
        "accept": "application/json",
        "X-Api-Key": JPI_API_KEY,
    }
    response = requests.get(JPI_RESOURCES_URL, headers=headers)
    assert response.ok, "Unable to fetch resources from JPI."
    resources = response.json()
    assert isinstance(resources, list), "Received invalid response from JPI."
    machines = list(filter(check_resource_group, resources))
    return machines


def convert_datetime(input: str) -> datetime:
    assert isinstance(input, str), "Invalid date and time string."
    return datetime.strptime(input, JPI_TIME_FORMAT)


def round_datetime(input: datetime) -> datetime:
    assert isinstance(input, datetime), "Invalid date and time data."
    if input.minute >= 30:
        return input.replace(minute=0, second=0) + timedelta(hours=1)
    else:
        return input.replace(minute=0, second=0)


def get_schedule_timeline(settings: dict[str, Any]) -> dict[str, datetime]:
    planning_start = convert_datetime(settings["PlanningStart"])
    days_before = timedelta(days=settings["DaysBeforePlanningStart"])
    planning_horizon = timedelta(weeks=settings["PlanningHorizon"])
    extra_day = timedelta(days=1)

    schedule_start = planning_start - days_before
    start_time = schedule_start.replace(hour=0, minute=0, second=0)

    schedule_end = planning_start + planning_horizon + extra_day
    end_time = schedule_end.replace(hour=0, minute=0, second=0)

    timeline = {"start_time": start_time, "end_time": end_time}

    return timeline


def iterate_timeline(timeline: dict[str, datetime]):
    current_time = timeline["start_time"]
    while current_time < timeline["end_time"]:
        yield current_time
        current_time += timedelta(hours=1)


def enumerate_timeline(timeline: dict[str, datetime]):
    return enumerate(iterate_timeline(timeline), 2)


def enumerate_headers():
    headers = ["Date", "Day", "Time"]
    return enumerate(headers, 1)


def enumerate_machines(machines: list[dict[str, Any]]):
    return enumerate(machines, 4)


def machine_index(resource: dict, machines: list[dict]) -> int:
    assert isinstance(resource, dict), "Invalid resource data."
    for col, machine in enumerate_machines(machines):
        if machine["Guid"] == resource["Guid"]:
            return col
    assert False


def check_resource_machine(resource: dict, machines: list[dict]) -> bool:
    assert isinstance(resource, dict), "Invalid resource data."
    assert isinstance(machines, list), "Invalid machines data."
    for machine in machines:
        assert isinstance(machine, dict)
        if machine["Guid"] == resource["Guid"]:
            return True
    return False


def other_resources(task: dict, machines: list[dict]):
    others = []
    resources = task["AssignedResources"]
    assert isinstance(resources, list), "Invalid resources data."
    for resource in resources:
        assert isinstance(resource, dict), "Invalid resource data."
        if not check_resource_machine(resource, machines):
            others.append(resource["Name"])
    return ", ".join(others)


def check_setup(task_name: str) -> bool:
    assert isinstance(task_name, str), "Invalid task name string."
    upper_name = task_name.upper()
    if "SETUP" in upper_name or "SET UP" in upper_name:
        return True
    return False


def chart(
    schedule: Worksheet,
    timeline: dict,
    machines: list[dict],
    job: dict,
    task: dict,
    resource: dict,
):
    assert isinstance(job, dict), "Invalid job data."
    assert isinstance(task, dict), "Invalid task data."
    assert isinstance(resource, dict), "Invalid resource data."
    assert isinstance(machines, list), "Invalid machines data."
    assert isinstance(timeline, dict), "Invalid timeline data."
    if not check_resource_machine(resource, machines):
        return
    col = machine_index(resource, machines)
    task_start = round_datetime(convert_datetime(task["Start"]))
    task_end = round_datetime(convert_datetime(task["End"]))
    count = 0
    for row, time in enumerate_timeline(timeline):
        if time < task_start or time >= task_end:
            continue
        count += 1
        setup = check_setup(task["Name"] or "")
        if count == 1:
            value = "Task: " + (task["Name"] or task["TaskNo"])
        elif count == 2:
            value = "Resin: " + (task["CustomFieldValue1"] or "N/A")
        elif count == 3 and setup:
            value = "Lines: " + (job["CustomFieldValue1"] or "N/A")
        elif time.hour == 7:
            value = "Continue: " + (task["Name"] or task["TaskNo"])
        else:
            value = "---"
        schedule.cell(row, col, value)
        if setup:
            schedule.cell(row, col).font = Font(bold=True)


def exception_offtimes(exception: dict[str, str]):
    workday = datetime.strptime(exception["Date"], JPI_TIME_FORMAT)
    nextday = workday + timedelta(days=1)
    worktimes = exception["WorkTime"].split(",")
    offtimes = []
    offtimes.append({"Start_Time": workday, "End_Time": None})
    for worktime in worktimes:
        if not worktime:
            continue
        start_time_str, end_time_str = worktime.split("-")
        start_time = datetime.strptime(start_time_str, "%H:%M").time()
        end_time = datetime.strptime(end_time_str, "%H:%M").time()
        worktime_start = datetime.combine(workday.date(), start_time)
        worktime_end = datetime.combine(workday.date(), end_time)
        offtimes[-1]["End_Time"] = worktime_start
        offtimes.append({"Start_Time": worktime_end, "End_Time": None})
    offtimes[-1]["End_Time"] = nextday
    return offtimes


def main():
    print("- Fetching system settings from JPI.")
    settings = get_jpi_settings()

    print("- Fetching jobs details from JPI.")
    jobs = get_active_jobs()

    print("- Fetching machine details from JPI.")
    machines = get_machine_resources()

    print("- Calculating schedule timeline.")
    time.sleep(1)

    timeline = get_schedule_timeline(settings)

    print("- Creating the Excel file.")
    time.sleep(1)

    workbook = Workbook()
    schedule = workbook.active
    assert isinstance(schedule, Worksheet), "Invalid Excel spreadsheet data."
    schedule.title = "Schedule"

    for col, header in enumerate_headers():
        schedule.cell(1, col, header)

    for col, machine in enumerate_machines(machines):
        schedule.cell(1, col, machine["Name"])

    for cell in schedule[1]:
        cell.border = Border(bottom=Side(style="medium"))
        cell.font = Font(bold=True)

    for row, input in enumerate_timeline(timeline):
        if input.hour == 7:
            schedule.cell(row, 1, input.strftime("%b %d")).font = Font(bold=True)
            schedule.cell(row, 2, input.strftime("%a")).font = Font(bold=True)
            for cell in schedule[row]:
                cell.border = Border(top=Side(style="medium"))
        schedule.cell(row, 3, input.strftime("%H")).font = Font(bold=True)

    schedule.freeze_panes = "D2"

    print("- Charting the machine schedule.")
    time.sleep(1)

    for job in jobs:
        assert isinstance(job, dict), "Invalid job data."
        tasks = job["Tasks"]
        assert isinstance(tasks, list), "Invalid tasks data."
        for task in tasks:
            assert isinstance(task, dict), "Invalid task data."
            resources = task["AssignedResources"]
            assert isinstance(resources, list), "Invalid resources data."
            for resource in resources:
                assert isinstance(resource, dict), "Invalid resource data."
                chart(schedule, timeline, machines, job, task, resource)

    color = PatternFill(start_color="BFB1D1", end_color="BFB1D1", fill_type="solid")

    for col, machine in enumerate_machines(machines):
        exceptions = machine["CalendarExceptions"]
        assert isinstance(
            exceptions, list
        ), "Invalid resource calendar exceptions data."
        for exception in exceptions:
            assert isinstance(
                exception, dict
            ), "Invalid resource calendar exception data."
            exception_date = convert_datetime(exception["Date"]).date()
            if exception_date < timeline["start_time"].date():
                continue
            if exception_date > timeline["end_time"].date():
                continue
            offtimes = exception_offtimes(exception)
            for row, input in enumerate_timeline(timeline):
                for offtime in offtimes:
                    start = round_datetime(offtime["Start_Time"])
                    end = round_datetime(offtime["End_Time"])
                    if input >= start and input < end:
                        schedule.cell(row, col).fill = color

    # Auto-fit the width of all columns
    for column in schedule.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = max_length + 2
        schedule.column_dimensions[
            get_column_letter(cell.column)
        ].width = adjusted_width

    print("- Saving the Excel file.")
    time.sleep(1)

    workbook.save(JPI_EXPORT_FILENAME)


if __name__ == "__main__":
    print(JPI_ASCII_ART)
    time.sleep(1)
    try:
        main()
    except AssertionError as error:
        print("Error:")
        if "JPI" in str(error):
            print("- A JPI related issue has occurred. Please try again later.")
        else:
            print("- Invalid data detected. Please try again later.")
        print("- If the issue persists, please contact your JPI representative.")
        input("Press Enter to exit...")
    except PermissionError:
        print("Error:")
        print(f"- The file {JPI_EXPORT_FILENAME} is currently open by another process")
        print("- Please close the file and try again.")
        input("Press Enter to exit...")
    except:
        print("Error:")
        print("- An unexpected error has occurred. Please try again later.")
        print("- If the issue persists, please contact your JPI representative.")
        input("Press Enter to exit...")
    else:
        print("- Schedule export completed successfully.")
        if platform.system() == "Windows":
            print("- Openning the Excel file.")
            subprocess.Popen(["start", " ", JPI_EXPORT_FILENAME], shell=True)
